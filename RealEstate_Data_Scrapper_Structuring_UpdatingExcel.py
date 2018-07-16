""" 
Created by Vytautas Bielinskas at Ekistics Real Estate
Last update: 2018 07 12

Definitions:
    ML - Machine learning
    BOW - Bag Of Words
"""

# :::-------------------------[Import libraries]----------------------------::: 
import requests, re, os
import pandas as pd
from bs4 import BeautifulSoup

from selenium import webdriver
options = webdriver.chrome.options.Options()
options.add_argument("--disable-extensions")
    
chrome_path = r"C:\...\chromedriver.exe"
driver = webdriver.Chrome(chrome_path)

# :::-----------------------------[Read BOW]--------------------------------:::
bow = pd.read_csv('bag_of_words.csv', header = None)
bow.insert(2, 'Counter', 0)

# :::---------------------[PROCEDURES AND FUNCTIONS]------------------------:::
# :::---------------------[Generate the list of URLs]-----------------------:::
def generateURLs(pages):
    listURLs = []
    base_url = "https://www.zoopla.co.uk/for-sale/property/station/rail/west-drayton/?identifier=station%2Frail%2Fwest-drayton&q=West Drayton Station%2C London&search_source=refine&radius=1&pn="
    for i in range(1, pages+1, 1):
        fullURL = base_url + str(i)
        listURLs.append(fullURL)
    return listURLs

# :::-------------------[Calculate Price per Bedroom]-----------------------:::
def calculatePricePerBedroom(DF):
    DF.insert(len(DF.columns), 'PRICE PER BEDROOM', 0)
    index_of_PPB = DF.columns.get_loc('PRICE PER BEDROOM')
    index_of_Bed = DF.columns.get_loc('BEDROOMS')
    index_of_Price = DF.columns.get_loc('CURRENT PRICE')
    
    for row in range(0, len(DF), 1):
        try:
            DF.iat[row, index_of_PPB] = float(DF.iat[row, index_of_Price]) / float(DF.iat[row, index_of_Bed])
        except:
            DF.iat[row, index_of_PPB] = 0.0
    return DF

# :::---------------[Visualize Evaluations of descriptions]-----------------:::
def visualizeEvaluationHist(df):
    import numpy as np
    import matplotlib.pyplot as plt
    
    bins = np.arange(0, max(DF['EVALUATION']), 0.05)
    
    _ = plt.hist(DF['EVALUATION'], bins, color = '#1f4e79')
    
    plt.xlabel('Evaluation of description')
    plt.ylabel('Number of properties')
    
    plt.show()
    return None

# :::-----------[Visualize Keyword Frequencies after Evaluation]------------:::
def visualizeKeywordsFreq(bow):
    import matplotlib.pyplot as plt
    import numpy as np
    
    bow = bow.sort_values('Counter', ascending = False)
    
    plt.style.use('_classic_test')
    
    my_color = '#5b9bd5'
    text_color = '#1f4e79'
    
    ax = bow[['Counter']].plot(kind='bar', title ="Keyword frequency",figsize=(7,4),legend=True, fontsize=11, color = my_color)
    
    ax.grid(color = '#cccccc', linestyle = '--', linewidth = 1)
    
    ax.set_xlabel("Keywords",fontsize=11, color = text_color)
    ax.set_ylabel("Frequency",fontsize=11, color = text_color)
    
    y_labels = np.arange(0, np.max(bow[['Counter']]), 20)
    ax.set_xticklabels(bow[0], rotation=90, fontsize = 11, color = text_color)
    ax.set_yticklabels(y_labels, fontsize = 11, color = text_color)
    
    return None

# :::-------------------------[Get all the data!]---------------------------:::
def getData(URLs, bow):
    
    size_counter = 0 # counting if size is written inside the property page
    
    properties = []
    
    # :::-----------------[Getting Data from internal page]-----------------:::
    def getDataFromInside(objectURL, size_counter, bow):
                
        # ::: ----------[Getting URL of Floorplan by Selenium]--------------:::
        def getFloorURL(objectURL):
                    
            driver.get(objectURL)
            driver.find_element_by_xpath(".//*[@id='ui-id-2']/span").click() # Click on Floorplan tab
            link = driver.find_element_by_partial_link_text("View origina").get_attribute("href")
                    
            return link

        # :::-------------------[Parse the description]---------------------:::
        def parseDescription(description):
            print('\nParsing description...')
            index_of_char = 0
            while description[index_of_char] == ' ':
                index_of_char = index_of_char + 1
            return description[index_of_char:]
        
        evaluation = 0.0
                
        print(objectURL)
        r = requests.get(objectURL)
        c = r.content
        soup = BeautifulSoup(c, "html.parser")
                
        '''
        
        IF DESIGN IS NOT CREEPY!!!!
        
        foundFloorPlan = False
                    
        ListOfTabs = soup.find_all("div", {"class":"ui-tabs"})
        NameOfTab = ListOfTabs[0].find_all("span")
        i = 0
        while len(NameOfTab[i].text.replace(" ", "").split("\n")[0]) > 1:
            if "Floorplan".upper() in NameOfTab[i].text.replace(" ", "").split("\n")[0].upper():
                foundFloorPlan = True
            #print("TAB ", i, ":", NameOfTab[i].text, " - ", foundFloorPlan)
            i = i + 1
        #print("--------------------")
                
        if foundFloorPlan == True:
            urlForFloorPlan = getFloorURL(objectURL)
        else:
            urlForFloorPlan = "Not found in server"
            
            '''
            
        url_block = str(soup.find('section', {'class':'dp-assets-card'})).replace(' ', '')
        
        try:
            urlForFloorPlan = url_block.split('<imgalt=""class=""src="')[1].split('"/')[0]
        except: 
            urlForFloorPlan = None
            
        print('FLOORPLAN URL:\n {}'.format(urlForFloorPlan), '\n \n')
        
        # :::--------[Getting features from the description inside]---------:::
        print('\nInside the: {}.'.format(objectURL))
        
        html = soup   # just for keep variables different for this part of def.
        
        max_sqft = 0.0
        this_sqft = 0.0
        
        # :::--------------[Get Size value from bullet text]----------------:::
        try:
            
            block_info = html.find('ul', {'class':'dp-features__list ui-list-bullets'})
            feature_list = block_info.find_all('li')
            feature_list_p = []
            size_list = len(feature_list)
            
            print('Yes, we found info block inside!\n{}\n'.format(str(block_info).replace('\n', '')))
            print('The size of the list is: {}.'.format(size_list))
            
            for feature_index in range(0, size_list, 1):
                feature_list_p.append(feature_list[feature_index].text)
                print('-- {}:{}.'.format(feature_index, feature_list[feature_index].text))
                
                if 'sq ft.'.upper().replace(' ', '').replace('.', '') in feature_list[feature_index].text.upper().replace(' ', '').replace('.', ''):
                    this_sqft = feature_list[feature_index].text.upper().replace(' ', '').replace('.', '').split('sq ft.'.upper().replace(' ', '').replace('.', ''))[0]
                    
                    print('\nThis sq ft.: {}'.format(this_sqft))
                    
                    if float(this_sqft) > max_sqft:
                        max_sqft = this_sqft
                        
                    size_counter = size_counter + 1
                    
                # :::---------[Evalute bullers features with BOW]-----------:::
                for word in range(0, len(bow), 1):
                    if bow.iat[word, 0].upper().replace(' ', '') in feature_list[feature_index].text.upper().replace(' ', ''):
                        evaluation = evaluation + float(bow.iat[word, 1])
                        print('-- keyword found: {}.'.format(bow.iat[word, 0]))
                
            #x = input('Go to the next property...')
            
        except:
            print('There are not bulltet info.\n')    
            
        # :::------------[Get description text and evaluate it]-------------:::
        description = html.find('div', {'class':'dp-description__text'}).text.replace('\n', '')
        description = parseDescription(description)
        print('\nDESCRIPTION: {}.'.format(description))
        
        # :::--------------[Evaluate this description by BOW]---------------:::
        description_temp = description.upper().replace(' ', '')
        
        for word in range(0, len(bow), 1):
            if bow.iat[word, 0].upper().replace(' ', '') in description_temp:
                evaluation = evaluation + float(bow.iat[word, 1])
                print('-- keyword found: {}.'.format(bow.iat[word, 0]))
                
                bow.iat[word, bow.columns.get_loc('Counter')] = bow.iat[word, bow.columns.get_loc('Counter')] + 1
            
        return urlForFloorPlan, max_sqft, size_counter, description, evaluation
    
    # :::------------------------[Get Stations data]------------------------:::
    def singleStation(station):
        station = station.text
        station = station.replace(' ', '')
        station = station.replace('\n', '')
        station = station.split('(')[1].split(')')[0]
        station = station.replace('miles', '')
        print('--> singleStation: {}'.format(station))
        return float(station)
    
    # :::-----------------------[Go through the pages]----------------------:::
    for page in range(0, len(URLs), 1):
        print("Page:", page)
        r = requests.get(URLs[page])
        c = r.content
        soup = BeautifulSoup(c, "html.parser")
        
        infoTable = soup.find("ul", {"class":"listing-results"})
        
        items = infoTable.find_all("div", {"class":"listing-results-wrapper"})
        for item in range(0, len(items), 1):
            blockProperty = items[item]    
            
            nearby_stations = blockProperty.find('div', {'class':'nearby_stations_schools'}).find_all('li')
            print('The size of stations: {}.'.format(len(nearby_stations)))
            
            number_of_stations = len(nearby_stations)
            
            # :::--------------[Build float array for distances]------------:::
            min_dist = 0   # Before evaluation
            try:
                stations_fl = []
                for this_station in range(0, len(nearby_stations), 1):
                    stations_fl.append(singleStation(nearby_stations[this_station]))
                min_dist = min(stations_fl)
            except:
                min_dist = -1
            
            print('MIN. DIST. {}.'. format(min_dist))
            
            """ get address : WORKING """
            address = blockProperty.find("a", {"class":"listing-results-address"}).text
            print(address)
            
            """ get area : WORKING """
            if "UB7" in address.upper():
                area_zone = "West Drayton"
            else:
                area_zone = ""
            
            """ get if object is new added : WORKING """
            attributes = blockProperty.find("h2").text
            new = False
            if "Just added".upper() in attributes.upper():
                new = True
            
            """ get listed on : WORKING"""
            listedOn = blockProperty.find("p", {"class":"top-half"}).find("small").text.split("\n")[2][1:]
            
            """ get Price : WORKING"""
            price = blockProperty.find("a", {"class":"text-price"}).text.replace(" ", "").split("\n")
            priceString = []
            for element in price:
                if element:
                    priceString.append(str(element))
            price = priceString[0].replace("£", "").replace(",", "")
            
            """ get First listed Price : WORKING"""
            sidebars = soup.find_all("div", {"class":"sidebar sbt"})
            firstListedSum = price
            for i in range(0, len(sidebars), 1):
                if "First listed".upper() in sidebars[i].text.upper():
                    firstListedSum = sidebars[i].text.split("First listed")[1].split(" on")[0].replace(" ", "").split("\n")
                    firstListedSum = firstListedSum[len(firstListedSum)-1].replace(",","").split("£")[1]
            
            """ get Bedrooms : WORKING """
            bedrooms = 0
            try:
                blockForAttrs = blockProperty.find("h3").find_all("span")
                for x in range(0, len(blockForAttrs), 1):
                    try:
                        if "num-beds" in blockForAttrs[x]["class"]:
                            bedrooms = float(blockForAttrs[x].text)
                    except:
                        print("No beds info")
            except:
                bedrooms = 0
            
            """ get Bathrooms : WORKING """
            bathrooms = 0
            try:
                blockForAttrs = blockProperty.find("h3").find_all("span")
                bathrooms = 0
                for x in range(0, len(blockForAttrs), 1):
                    try:
                        if "num-baths" in blockForAttrs[x]["class"]:
                            bathrooms = float(blockForAttrs[x].text)
                            print('BATHROOMS', bathrooms)
                    except:
                        print("No bath info")
            except:
                bathrooms = 0
            #print(bathrooms)
            
            """ get Area : WORKING """
            area = 0
            try:
                blockForAttrs = blockProperty.find("h3").find_all("span")
                for x in range(0, len(blockForAttrs), 1):
                    if "num-sqft" in blockForAttrs[x]["class"]:
                        try:
                            area = float(blockForAttrs[x].text.split(" ")[0].replace(",",""))
                            print('AREA:', area)
                        except:
                            print("No area")
            except:
                print(area)
            
            """ get Reception rooms : WORKING """
            receptions = 0
            try:
                blockForAttrs = blockProperty.find("h3").find_all("span")
                receptions = 0
                for x in range(0, len(blockForAttrs), 1):
                    try:
                        if "num-reception" in blockForAttrs[x]["class"]:
                            receptions = float(blockForAttrs[x].text)
                    except:
                        print("No bath")
            except:
                receptions = 0
            #print(receptions) 
            
            """ get link : WORKING """
            linkTo = str(blockProperty.find("a")).split('"/')[1].split('"')[0]
            linkTo = "https://www.zoopla.co.uk/" + linkTo
            #print(linkTo)
            
            """ get ID : WORKING """
            propID = blockProperty.parent["data-listing-id"]
            #print(propID)
            
            # :::--------[Get all possible data from internal page]---------:::
            data_inside = getDataFromInside(linkTo, size_counter, bow)
            FloorPlanURL = data_inside[0]
            size_counter = data_inside[2]
            description = data_inside[3]
            evaluation = data_inside[4]
            print('\nSize counter: {}.'.format(size_counter))
            # ::: ----------------------------------------------------------:::
            
            if float(data_inside[1]) > float(area):
                area = data_inside[1]
            
            """ Save features to Database """
            features = {}
            if new == True:
                features["NEW"] = "New"
            else:
                features["NEW"] = ""
                
            features["ID"] = propID
            features["AREA"] = area_zone
            features["FIRST LISTED"] = listedOn
            features["ADDRESS"] = address
            features['NUMBER OF CLOSEST STATIONS'] = number_of_stations
            features['MIN. DISTANCE TO THE STATION'] = min_dist
            features["SIZE"] = area
            features["CURRENT PRICE"] = price
            features["ORIGINAL PRICE"] = firstListedSum
            features["BATHROOMS"] = bathrooms
            features["BEDROOMS"] = bedrooms
            features["RECEPTION"] = receptions
            features["FLOORPLAN LINK"] = FloorPlanURL
            features['DESCRIPTION'] = description
            features["WEB"] = linkTo
            features['EVALUATION'] = evaluation
            
            print(features, '\n')
                
            properties.append(dict(features))            
            
    return properties, size_counter

# :::-------------------------[UPDATING THE EXCEL FILE]---------------------:::
""" 
            All new data will be written to File_before.xlsx Excel file
            
"""
def updateExcel(dataset):
    
    print('# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #')
    print('Writing data to Excel file...')
    
    import requests, re, os
    import pandas as pd
    from bs4 import BeautifulSoup

    import openpyxl, os, datetime
    from pandas import ExcelWriter as ewriter
    from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Fill
    from openpyxl.comments import Comment
    
    """ checking if property is NOT LAND : start """
    def ifNotLand(url):
        land = False
        
        r = requests.get(url)
        c = r.content
        soup = BeautifulSoup(c, "html.parser")
        
        title = soup.find("title").text
        if "land for sale".upper() in title.upper():
            land = True
        return land
    """ checking if property is NOT LAND : end """
    
    
    cwd = os.getcwd()
    directory = (cwd)                                           #<-- Change direcotry to Excel folder
    os.chdir(directory)
    os.getcwd()
    
    filename = "File_before.xlsx"
    wb = openpyxl.load_workbook(filename, data_only = False)    
    
    #print(wb.get_sheet_names)
    print("---- Start update Excel file ----")       
    
    sheet = wb.get_sheet_by_name("Data Sales")
    
    rowForSearching = 10
    while len(str(sheet.cell(row = rowForSearching, column = 2).value)) > 4:
        rowForSearching = rowForSearching + 1
        
    end_of_table = rowForSearching - 1                          # Get the end of current table
    
    """ Define some Excel cell styles and formatting """
    fillDefault = PatternFill(fill_type = None,
                              start_color = "FFFFFFFF",
                              end_color = "FF000000")
    
    fillGreen = PatternFill(fill_type = "solid",
                            start_color = "56af11",
                            end_color = "56af11")
    
    HyperlinkBlue = Font(color = "0563c1",
                         underline = "single")
    
    rightAligment = Alignment(horizontal = "right")
    centerAligment = Alignment(horizontal = "center")
    
    """ Starting writing data to Excel file """
    id_column      = dataset.columns.get_loc("ID")            # get index of ID column
    id_area        = dataset.columns.get_loc("AREA")          # get index of STATION column
    id_address     = dataset.columns.get_loc("ADDRESS")       # get index of ADDRESS column
    id_bedrooms    = dataset.columns.get_loc("BEDROOMS")      # get index of BEDROOMS column
    id_bathrooms   = dataset.columns.get_loc("BATHROOMS")     # get index of BATHROOMS column
    id_reception   = dataset.columns.get_loc("RECEPTION")     # get index of RECEPTION column
    id_size        = dataset.columns.get_loc("SIZE")          # get index of SIZE column
    id_firstListed = dataset.columns.get_loc("FIRST LISTED")  # get index of FIRST LISTED column
    id_origPrice   = dataset.columns.get_loc("ORIGINAL PRICE")# get index of ORIGINAL PRICE column
    id_currPrice   = dataset.columns.get_loc("CURRENT PRICE") # get index of CURRENT PRICE column
    id_link        = dataset.columns.get_loc("WEB")           # get index of WEB column
    id_floorplan   = dataset.columns.get_loc("FLOORPLAN LINK")# get index of FLOORPLAN LINK column
    id_new         = dataset.columns.get_loc("NEW")           # get index of NEW column
    id_evaluation  = dataset.columns.get_loc('EVALUATION')    # get index of EVALUATION column
    
    for x in range(rowForSearching, rowForSearching + len(dataset), 1):
        for j in range(1, sheet.max_column, 1):
            sheet.cell(row = x, column = j).fill = fillDefault
    
    dataset_index = 0
    property_index = rowForSearching
    for i in range(rowForSearching, rowForSearching + len(dataset), 1):
        link_to_property = dataset.iat[dataset_index, id_link]
        land = ifNotLand(link_to_property)
        duplicates = 0
        if land == False and "POA" not in str(dataset.iat[dataset_index, id_currPrice]).upper():
            """ writing MONTH column """
            d = datetime.datetime.strptime(str(datetime.date.today()), '%Y-%m-%d')
            d= d.strftime('%d-%b-%Y').replace("-20", "-")
            sheet.cell(row = property_index, column = 2).value = d
            sheet.cell(row = property_index, column = 2).alignment = rightAligment
            
            """ writing SIZE columns """
            size = dataset.iat[dataset_index, id_size]
            if str(size).replace(" ", "") == "0":
                sheet.cell(row = property_index, column = 9).value = ""
            else:
                sheet.cell(row = property_index, column = 9).value = dataset.iat[dataset_index, id_size]
            
            """ writing ORIGINAL PRICE columns """
            sheet.cell(row = property_index, column = 11).value = dataset.iat[dataset_index, id_origPrice]
            sheet.cell(row = property_index, column = 11).alignment = rightAligment
            
            """ writing ID column """
            sheet.cell(row = property_index, column = 3).value = dataset.iat[dataset_index, id_column]
            sheet.cell(row = property_index, column = 3).alignment = rightAligment
            
            """ checking past data by ID : start --> """
            for past_row in range(10, property_index-1, 1):
                past_id = str(sheet.cell(row = past_row, column = 3).value)
                if str(sheet.cell(row = property_index, column = 3).value) == past_id:
                    duplicates = duplicates + 1
                    sheet.cell(row = property_index, column = 3).fill = fillGreen
                    
                    formula_for_comps = '=IF(AND(F' + str(property_index) + '<>2, F' + str(property_index) + '<>4), "",IFERROR(VLOOKUP(C' + str(property_index) + ',$C$3:$V$' + str(end_of_table) + ',20,FALSE), "new"))'
                    
                    size = sheet.cell(row = past_row, column = 9).value
                    sheet.cell(row = property_index, column = 9).value = size
                    sheet.cell(row = property_index, column = 9).fill = fillGreen
                    
                    print("Duplicate on row:", past_row, ", new property is on", property_index, "row")
            """ checking past data by ID : end --> """
            
            """ writing AREA column """
            sheet.cell(row = property_index, column = 4).value = dataset.iat[dataset_index, id_area]
            
            """ writing ADDRESS column """
            sheet.cell(row = property_index, column = 5).value = dataset.iat[dataset_index, id_address]
            
            """ writing BEDROOMS column """
            sheet.cell(row = property_index, column = 6).value = dataset.iat[dataset_index, id_bedrooms]
            
            """ writing BATHROOMS columns """
            sheet.cell(row = property_index, column = 7).value = dataset.iat[dataset_index, id_bathrooms]
            
            """ writing RECEPTION columns """
            sheet.cell(row = property_index, column = 8).value = dataset.iat[dataset_index, id_reception]
            
            """ writing FIRST LISTED columns """
            sheet.cell(row = property_index, column = 10).value = dataset.iat[dataset_index, id_firstListed]
            sheet.cell(row = property_index, column = 10).alignment = rightAligment
            
            """ writing FLOOR PLAN LINK columns """
            FullLink = ""
            print(':--> {}'.format(dataset.iat[dataset_index, id_floorplan]))
            try:
                if ".com" in dataset.iat[dataset_index, id_floorplan] != None:
                    if ".com" in dataset.iat[dataset_index, id_floorplan]:
                        FullLink = '=HYPERLINK("' + dataset.iat[dataset_index, id_floorplan] + '","' + "Floor" + '")'
                        sheet.cell(row = property_index, column = 15).font = HyperlinkBlue
                    else:
                        FullLink = ""
                    sheet.cell(row = property_index, column = 15).value = FullLink
                    sheet.cell(row = property_index, column = 15).alignment = centerAligment
            except:
                print('\nWriting Floorplan was not succifient.')
            
            """ writing CURRENT PRICE to columns """
            sheet.cell(row = property_index, column = 12).value = dataset.iat[dataset_index, id_currPrice]
            sheet.cell(row = property_index, column = 12).alignment = rightAligment
            
            """ writing LINK to column """
            FullLink = '=HYPERLINK("' + dataset.iat[dataset_index, id_link] + '","' + "Link" + '")'
            sheet.cell(row = property_index, column = 16).value = FullLink
            sheet.cell(row = property_index, column = 16).font = HyperlinkBlue
            
            """ writing LINK URL to column """
            sheet.cell(row = property_index, column = 18).value = dataset.iat[dataset_index, id_link]
            
            """ writing NEW columns """
            sheet.cell(row = property_index, column = 21).value = dataset.iat[dataset_index, id_new]
            
            """ writing EVALUATION column """
            sheet.cell(row = property_index, column = 28).value = dataset.iat[dataset_index, id_evaluation]
            sheet.cell(row = property_index, column = 28).fill = fillDefault
            
            dataset_index = dataset_index + 1           #Go to next row
            property_index = property_index + 1
            
            """ Write formulas : start """
            formula_for_comps = '=IF(AND(F' + str(property_index) + '<>2, F' + str(property_index) + '<>4), "",IFERROR(VLOOKUP(C' + str(property_index) + ',$C$3:$V$' + str(end_of_table) + ',20,FALSE), "new"))'
                                        
            sheet.cell(row = property_index, column = 27).value = formula_for_comps
            sheet.cell(row = property_index, column = 27).fill = fillGreen
                    
            formula_check_size = '=IF(AND(Y' + str(property_index) + '="new", I' + str(property_index) + '=""), "check", IF(I' + str(property_index) + '="", Y' + str(property_index) + ',I' + str(property_index) + '))'
            sheet.cell(row = property_index, column = 26).value = formula_check_size
            sheet.cell(row = property_index, column = 26).fill = fillGreen
            """ Write formulas : end """
            
            """ Calculation: DISCOUNT column """
            sheet.cell(row = property_index, column = 13).value = '=IFERROR((L' + str(property_index) + '-K' + str(property_index) + ')/K' + str(property_index)+ ',"")'
            sheet.cell(row = property_index, column = 13).number_format = "0%"
            
            """ Calculation: PSF column """
            sheet.cell(row = property_index, column = 14).value = '=IFERROR(L' + str(property_index) + '/I' + str(property_index) + ',"")'
            
            """ Calculation: Type column """
            sheet.cell(row = property_index, column = 24).value = '=IF(F' + str(property_index) + '=0,"Studio",IF(F' + str(property_index) + '>=5,"5+",F' + str(property_index) + '))'        
            
        else:
            print("Row", i, "- this is LAND FOR SALE")
            dataset_index = dataset_index + 1
            
    """ SAVE DATA TO EXCEL """
    import datetime
    from datetime import date 
    timestamp = str(datetime.datetime.now()).split('.')[0].replace('-', '').replace(':', '').replace(' ', '_')
    
    area_values = ['West Drayton', 
                   'Ilford']
    area_value = area_values[0]
    
    wb.save(timestamp + " File_After_SALES - " + area_value + ".xlsx")
    print('Finished.')
    
    return None    

# :::------------------------[Structuring my dataframe]---------------------:::
# :::---------------------------[Generate URL list]-------------------------:::
number_of_pages = 9
URLs = generateURLs(number_of_pages)

# :::--------------------------------[MAIN PART]----------------------------:::
MyBigData = getData(URLs, bow)
FullData = MyBigData[0]

DF = pd.DataFrame(FullData)
DF = DF[['ID', 'AREA', 'ADDRESS', 'BEDROOMS', 'BATHROOMS', 'RECEPTION', 'SIZE', 
         "FIRST LISTED", 'ORIGINAL PRICE', 'CURRENT PRICE', 'WEB', 'FLOORPLAN LINK', 
         "NEW", 'NUMBER OF CLOSEST STATIONS', 'MIN. DISTANCE TO THE STATION',
         'EVALUATION']]

# ::: -----[Sort values from the highest evaluation to the lowest one]------:::
DF = DF.sort_values(by = ['EVALUATION', 'MIN. DISTANCE TO THE STATION'], 
               ascending = False)

DF = calculatePricePerBedroom(DF)

DataBase = DF

size_counter = MyBigData[1]
print('We have {} properties with size value inside property page.'.format(size_counter))

# ::: -------[Convert negatives values to the representative ones]--------- :::
max_distance = max(DataBase['MIN. DISTANCE TO THE STATION'])
theor_max_d  = max_distance * 1.5

index_of_st_dist = DataBase.columns.get_loc('MIN. DISTANCE TO THE STATION')
for this_row in range(0, len(DataBase), 1):
    if DataBase.iat[this_row, index_of_st_dist] == float(-1):
        DataBase.iat[this_row, index_of_st_dist] = theor_max_d
        
DF_analyze = DataBase.drop(columns=['NUMBER OF CLOSEST STATIONS',
                            'MIN. DISTANCE TO THE STATION']
    )
    
updateExcel(DF_analyze)
# :::------------------------[Data Visualization]---------------------------:::
visualizeEvaluationHist(DF)
visualizeKeywordsFreq(bow)

# :::------- [Write data to CSV file to current working directory] ---------:::
import datetime
cwd = os.getcwd()
os.chdir(cwd)
timestamp = str(datetime.datetime.now()).split('.')[0].replace('-', '').replace(':', '').replace(' ', '_')
DataBase.to_csv(timestamp + '_Zoopla.co.uk_WESTDRAYTON_SALES.csv')
# :::-----------------------------------------------------------------------:::
